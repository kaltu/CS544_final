from openpyxl import Workbook, load_workbook
import psycopg2
import sys

# 0. Define all the variable
book = load_workbook('notesResult.xlsx')
sh = book['Results']
hostname = 'localhost'
database = 'mimic3'
username = 'postgres'
pwd = 'password'
port_id = 5432
l1 = []
l2 = []

# 1. connect to the database
try:
    with psycopg2.connect(
        host = hostname,
        dbname = database,
        user = username,
        password = pwd,
        port = port_id) as conn:

        # create a cursor to execute SQL code
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM noteevents")


            for row in cur:
                l1.append(row)

            for i in range(len(l1)):
                text = l1[i][10]
                words = text.split()

                if 'sepsis' in words and (words[words.index('sepsis')-1] != 'not' or words[words.index('sepsis')-1] != 'no'):
                    l2.append(l1[i])                        

                elif 'septic' in words and words[words.index('septic')-1] != 'not':                        
                    l2.append(l1[i])

            # write the header
            sh.cell(1,1).value = 'row_id'
            sh.cell(1,2).value = 'subject_id'
            sh.cell(1,3).value = 'hadm_id'
            sh.cell(1,4).value = 'chartdate'
            sh.cell(1,5).value = 'charttime'
            sh.cell(1,6).value = 'storetime'
            sh.cell(1,7).value = 'category'
            sh.cell(1,8).value = 'description'
            sh.cell(1,9).value = 'cgid'
            sh.cell(1,10).value = 'iserror'
            sh.cell(1,11).value = 'text'

            # write data from l2 into sh
            for i in range(len(l2)):
                for j in range(11):
                    sh.cell(i + 2, j + 1).value = l2[i][j]

            book.save('notesResult.xlsx')

except Exception as error:
    print(error)
finally:
    if conn is not None:
        conn.close()